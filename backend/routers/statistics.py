from fastapi import APIRouter, Depends, Response
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import distinct
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from io import BytesIO
import datetime

from database import get_db
from models import Reservation, User, Room

router = APIRouter(
    prefix="/api/report",
    tags=["report"]
)

@router.get("/xlsx")
def export_all_reservations(db: Session = Depends(get_db)):
    try:
        time_slots = [
            "08:00", "09:45", "11:30", "13:15", "15:00", "16:45", "18:30",
        ]

        rooms = db.query(Room).all()
        dates = db.query(distinct(Reservation.date)).order_by(Reservation.date).all()
        dates = [d[0] for d in dates if d[0] is not None]
        reservations = db.query(Reservation).options(joinedload(Reservation.user)).all()

        wb = Workbook()
        ws = wb.active
        ws.title = "Harmonogram"

        row_idx = 1

        for day in dates:
            day_str = day.strftime("%A %d.%B").capitalize()

            # Dzień - scalona komórka
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2 + len(time_slots))
            ws.cell(row=row_idx, column=1, value=f"DZIEŃ: {day_str}")
            ws.cell(row=row_idx, column=1).font = Font(bold=True)
            row_idx += 1

            # Godziny
            ws.cell(row=row_idx, column=1, value="SALA").font = Font(bold=True)
            for i, slot in enumerate(time_slots):
                ws.cell(row=row_idx, column=i + 2, value=slot).font = Font(bold=True)
            row_idx += 1

            for room in rooms:
                ws.cell(row=row_idx, column=1, value=room.name)

                for i, slot_start_str in enumerate(time_slots):
                    slot_start = datetime.datetime.strptime(slot_start_str, "%H:%M").time()
                    slot_end = (datetime.datetime.combine(datetime.date.today(), slot_start) + datetime.timedelta(minutes=90)).time()

                    reservation = next((
                        r for r in reservations
                        if r.date == day and r.room_id == room.id
                        and r.time_from and r.time_to
                        and r.time_from < slot_end and r.time_to > slot_start
                    ), None)

                    if reservation and reservation.user and reservation.user.first_name and reservation.user.last_name:
                        user_info = f"{reservation.user.first_name[0]}.{reservation.user.last_name}"
                    else:
                        user_info = ""

                    cell = ws.cell(row=row_idx, column=i + 2, value=user_info)
                    cell.alignment = Alignment(horizontal="center")

                row_idx += 1

            row_idx += 2

        # Auto-szerokość (z poprawką na MergedCell)
        for i, col in enumerate(ws.columns, start=1):
            max_length = max(len(str(cell.value or "")) for cell in col)
            col_letter = get_column_letter(i)
            ws.column_dimensions[col_letter].width = max_length + 2

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        headers = {
            "Content-Disposition": "attachment; filename=harmonogram_formatowany.xlsx"
        }

        return Response(
            content=output.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers
        )

    except Exception as e:
        import traceback
        print("Błąd eksportu XLSX:", e)
        traceback.print_exc()
        return Response(content=f"Błąd eksportu XLSX: {e}", status_code=500)
